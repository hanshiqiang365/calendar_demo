#author: hanshiqiang365 （微信公众号：韩思工作室）

from openpyxl.styles import Alignment, PatternFill, Font
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
import openpyxl
import calendar
import numpy as np


calendar.setfirstweekday(firstweekday=6)

wb = openpyxl.Workbook()
align = Alignment(horizontal='center', vertical='center')
    
for i in range(1, 13):
    sheet = wb.create_sheet(index=0, title=str(i) + '月')
    for j in range(len(calendar.monthcalendar(2020, i))):
        for k in range(len(calendar.monthcalendar(2020, i)[j])):
            value = calendar.monthcalendar(2020, i)[j][k]
            if value == 0:
                value = ''
                sheet.cell(row=j + 5, column=k + 1).value = value
                sheet.cell(row=j + 5, column=k + 1).alignment = align
            else:
                sheet.cell(row=j + 5, column=k + 1).value = value
                sheet.cell(row=j + 5, column=k + 1).font = Font(u'微软雅黑', size=11)
                sheet.cell(row=j + 5, column=k + 1).alignment = align


    fill = PatternFill("solid", fgColor="B9EBF7")
    for k1 in range(1, 100):
        for k2 in range(1, 100):
            sheet.cell(row=k1, column=k2).fill = fill
    days = ['星期日', '星期一', '星期二', '星期三', '星期四', '星期五', '星期六']
    num = 0
    
    for k3 in range(1, 8):
        sheet.cell(row=4, column=k3).value = days[num]
        sheet.cell(row=4, column=k3).alignment = align
        sheet.cell(row=4, column=k3).font = Font(u'微软雅黑', size=11)
        c_char = get_column_letter(k3)
        sheet.column_dimensions[get_column_letter(k3)].width = 12
        num += 1
        
    sheet.merge_cells('A11:G25')
    img = Image(f'hanshiqiang365.png')
    sheet.add_image(img, 'A11')

    sheet.cell(row=2, column=1).value = '2020年'
    sheet.cell(row=2, column=2).value = str(i) + '月'

    sheet.cell(row=2, column=1).font = Font(u'微软雅黑', size=16, bold=True, color='FF7887')
    sheet.cell(row=2, column=2).font = Font(u'微软雅黑', size=16, bold=True, color='FF7887')
    sheet.cell(row=2, column=1).alignment = align
    sheet.cell(row=2, column=2).alignment = align

wb.save('2020calendar.xlsx')
